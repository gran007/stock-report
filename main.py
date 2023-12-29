import openpyxl
import os
import glob

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from db_manager import *


def insert_excel_data():
    company_info = {}

    input_path = 'C:\\Users\\gran007\\Desktop\\주식\\202311'
    for file_path in glob.glob(f'{input_path}/*.xlsx'):
        basename = os.path.basename(file_path)
        date = os.path.splitext(basename)[0]
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        delete_daily_data(date)
        for row in [r for r in ws.rows][1:]:
            row_data = [c.value for c in row]
            company_info[row_data[0]] = row_data[1]
            insert_daily_data(date, row_data)
        commit()

    for id, name in company_info.items():
        insert_company_type(id, name)
    commit()


headers = ['종목코드', '종목명', '거래량초과']


def export_excel(dates, result):
    output_path = 'C:\\Users\\gran007\\Desktop\\주식\\12월 종합.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 10

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    align_center = Alignment(vertical='center', horizontal='center')
    align_right = Alignment(vertical='center', horizontal='right', wrapText=True)
    header_bg_color = PatternFill(start_color='DBE5F1', end_color='DBE5F1', fill_type='solid')

    for index, name in enumerate(headers):
        cell = ws.cell(1, index + 1)
        cell.value = name
        cell.fill = header_bg_color
        cell.alignment = align_center
        cell.border = thin_border
    for index, date in enumerate(dates):
        cell = ws.cell(1, len(headers) + index + 1)
        cell.value = date.strftime('%m월 %d일')
        cell.fill = header_bg_color
        ws.column_dimensions[chr(68+index)].width = 12
        cell.alignment = align_center
        cell.border = thin_border

    for row_index, (code, item) in enumerate(result):
        ws.cell(row_index + 2, 1).value = code
        ws.cell(row_index + 2, 1).font = Font(bold=True)
        ws.cell(row_index + 2, 1).alignment = align_center
        ws.cell(row_index + 2, 2).value = item['name']
        ws.cell(row_index + 2, 2).alignment = align_center
        ws.cell(row_index + 2, 3).value = f"{item['count']}회"
        ws.cell(row_index + 2, 3).alignment = align_right
        ws.row_dimensions[row_index + 2].height = 52

        for col_index, date in enumerate(dates):
            if date in item:
                row_item = item[date]
                fluctuation_ratio = round(row_item['fluctuation_ratio'], 2)
                closing_price = format(row_item['closing_price'], ',')
                is_over_transaction_amount_standard = row_item['is_over_transaction_amount_standard']
                cell = ws.cell(row_index + 2, col_index + 4)
                cell.alignment = align_right
                cell.value = f'="{fluctuation_ratio}%"&CHAR(10)&"({closing_price}원)"'
                cell.border = thin_border
                if fluctuation_ratio >= 0:
                    bg_color = 'E6B9B8' if is_over_transaction_amount_standard else 'F2DDDC'
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    cell.font = Font(color='FD0012', bold=True if is_over_transaction_amount_standard else False)
                else:
                    bg_color = '4F81BD' if is_over_transaction_amount_standard else 'B8CCE4'
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    cell.font = Font(color='38497D', bold=True if is_over_transaction_amount_standard else False)
    wb.save(output_path)


if __name__ == '__main__':
    dates, result = select_date_and_transaction_amount('2023-12', 200_000_000_000)
    export_excel(dates, result)

